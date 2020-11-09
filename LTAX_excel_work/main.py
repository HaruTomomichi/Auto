import os
import openpyxl

class AUTO_SYSTEM():
    def __init__(self):
        # 첫번째 원소는 t / 두번째 원소는 e입니다
        self.a_sum = [0, 0]
        self.b_sum = [0, 0]
        self.c_sum = [0, 0]
        self.d_sum = [0, 0]
        self.e_sum = [0, 0]
        self.f_sum = [0, 0]
        self.g_sum = [0, 0]
        self.r_sum = 0

        self.output = openpyxl.load_workbook(os.getcwd() + '/LTAX_값진단종합결과_테이블별오류취합_템플릿_200915_v0.01.xlsx')['Sheet1']

    def main_action(self):

        while 1:
            f_name = input("file name : ")

            file_name = os.getcwd() + '/' + f_name
            wb = openpyxl.load_workbook(file_name)
            sheet = wb['(진단실행)진단항목오류정보']

            table = sheet.cell(row=2, column=1).value

            self.second_action()
            self.fourth_action(f_name)

            print('Done')
            check = input('continue? (y/n): ')
            if (check == 'n'):
                break

    def second_action(self):

    def third_action(self,column_val):
        if (column_val == '날짜'):
            self.a_sum[0] = self.total_sum(self.a_sum[0])
            self.a_sum[1] = self.error_sum(self.a_sum[1])
            if (self.a_sum[1] == 0):
                self.r_sum = 0
            else:
                self.r_sum = float((self.a_sum[1] / self.a_sum[0]) * 100)
            self.write(2, self.a_sum[0], self.a_sum[1])

        elif (column_val == '여부'):
            self.b_sum[0] = self.total_sum(self.b_sum[0])
            self.b_sum[1] = self.error_sum(self.b_sum[1])
            if (self.b_sum[1] == 0):
                self.r_sum = 0
            else:
                self.r_sum = float((self.b_sum[1] / self.b_sum[0]) * 100)
            self.write(2, self.b_sum[0], self.b_sum[1])

        elif (column_val == '번호'):
            self.c_sum[0] = self.total_sum(self.c_sum[0])
            self.c_sum[1] = self.error_sum(self.c_sum[1])
            if (self.c_sum[1] == 0):
                self.r_sum = 0
            else:
                self.r_sum = float((self.c_sum[1] / self.c_sum[0]) * 100)
            self.write(2, self.c_sum[0], self.c_sum[1])

        elif (column_val == '금액'):
            self.d_sum[0] = self.total_sum(self.d_sum[0])
            self.d_sum[1] = self.error_sum(self.d_sum[1])
            if (self.d_sum[1] == 0):
                self.r_sum = 0
            else:
                self.r_sum = float((self.d_sum[1] / self.d_sum[0]) * 100)
            self.write(2, self.d_sum[0], self.d_sum[1])

        elif (column_val == '수량'):
            self.e_sum[0] = self.total_sum(self.e_sum[0])
            self.e_sum[1] = self.error_sum(self.e_sum[1])
            if (self.e_sum[1] == 0):
                self.r_sum = 0
            else:
                self.r_sum = float((self.e_sum[1] / self.e_sum[0]) * 100)
            self.write(2, self.e_sum[0], self.e_sum[1])

        elif (column_val == '율'):
            self.f_sum[0] = self.total_sum(self.f_sum[0])
            self.f_sum[1] = self.error_sum(self.f_sum[1])
            if (self.f_sum[1] == 0):
                self.r_sum = 0
            else:
                self.r_sum = float((self.f_sum[1] / self.f_sum[0]) * 100)
            self.write(2, self.f_sum[0], self.f_sum[1])

        elif (column_val == '코드'):
            self.g_sum[0] = self.total_sum(self.g_sum[0])
            self.g_sum[1] = self.error_sum(self.g_sum[1])
            if (self.g_sum[1] == 0):
                self.r_sum = 0
            else:
                self.r_sum = float((self.g_sum[1] / self.g_sum[0]) * 100)
            self.write(2, self.g_sum[0], self.g_sum[1])

    def total_sum(self,t_sum):
        total = openpyxl.load_workbook(file_name)[['(진단실행)진단항목오류정보'].cell(row=i, column=5).value
        while (total.find(',') > 0):
            total = total.replace(',', '')
        t_sum += int(total)
        return t_sum

    def error_sum(self,e_sum):
        pass

    def write(self,domain,t_sum,e_sum,r_sum):
        self.output.cell(row=j, column=domain, value=t_sum)
        self.output.cell(row=j, column=domain + 1, value=e_sum)
        self.output.cell(row=j, column=domain + 2, value=self.r_sum)

    def fourth_action(self,f_name):
        self.output.cell(row=j, column=1, value='합계')

        for col in range (2, 23) :
            result = 0.0
            for row in range (4, j) :
                if hasattr(self.output.cell(row=row, column=col), 'value') :
                    if self.output.cell(row=row, column=col).value is not None :
                        self.user_sum += self.output.cell(row=row, column=col).value
                    else :
                        self.output.cell(row=j, column=col, value = 0)
            self.output.cell(row=j, column=col, value=result)

        f_name = os.getcwd() + '/test/output_' + f_name
        openpyxl.load_workbook(os.getcwd()+'/LTAX_값진단종합결과_테이블별오류취합_템플릿_200915_v0.01.xlsx').save(f_name)

    def user_sum(self,):
        temp = 0
        for row in range(4, j):
            temp += self.output.cell(row=row, column=col).value
        return sum




if __name__=="__main__" :

    on_work = AUTO_SYSTEM()
    on_work.main_action()