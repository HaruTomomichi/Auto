import os

import openpyxl

global at_sum, bt_sum, ct_sum, dt_sum, et_sum, ft_sum, gt_sum
global ae_sum, be_sum, ce_sum, de_sum, ee_sum, fe_sum, ge_sum
global r_sum


# 전체건수
def total_sum(t_sum) :
    total = sheet.cell(row=i, column=5).value
    while (total.find(',') > 0):
        total = total.replace(',', '')
    t_sum += int(total)
    return t_sum

# 오류건수
def error_sum(e_sum) :
    error = sheet.cell(row=i, column=6).value
    while (error.find(',') > 0):
        error = error.replace(',', '')
    e_sum += int(error)
    return e_sum

# output
def write(domain, t_sum, e_sum) :
    output.cell(row=j, column=domain, value=t_sum)
    output.cell(row=j, column=domain+1, value=e_sum)
    output.cell(row=j, column=domain+2, value=r_sum)

def sum(col, j) :
    sum = 0
    for row in range (4,j) :
        sum += output.cell(row=row, column=col).value
    return sum

if __name__=="__main__" :

    while True :
        f_name = input("file name : ")
        cur_dir = os.getcwd()
        file_name = cur_dir+ '/test/' + f_name
        wb = openpyxl.load_workbook(file_name)
        sheet = wb['(진단실행)진단항목오류정보']

        out = openpyxl.load_workbook(cur_dir+'/LTAX_값진단종합결과_테이블별오류취합_템플릿_200915_v0.01.xlsx')
        output = out['Sheet1']

        flag = True
        i = 2
        j = 4

        at_sum = bt_sum = ct_sum = dt_sum = et_sum = ft_sum = gt_sum = 0
        ae_sum = be_sum = ce_sum = de_sum = ee_sum = fe_sum = ge_sum = 0

        table = sheet.cell(row=2, column=1).value

        while flag :
            t_read = sheet.cell(row=i, column=1).value
            if (t_read == table) :
                output.cell(row=j, column=1, value=table)

            else:
                if (t_read == None):
                    flag = False
                j = j+1
                table = t_read
                output.cell(row=j, column=1, value=table)
                at_sum = bt_sum = ct_sum = dt_sum = et_sum = ft_sum = gt_sum = 0
                ae_sum = be_sum = ce_sum = de_sum = ee_sum = fe_sum = ge_sum = 0
                r_sum = 0

            d_read = sheet.cell(row=i, column=3).value

            if (d_read == '날짜'):
                at_sum = total_sum(at_sum)
                ae_sum = error_sum(ae_sum)
                if (ae_sum == 0) :
                    r_sum = 0
                else :
                    r_sum = float((ae_sum/at_sum)*100)
                write(2, at_sum, ae_sum)

            if (d_read == '여부'):
                bt_sum = total_sum(bt_sum)
                be_sum = error_sum(be_sum)
                if (be_sum == 0):
                    r_sum = 0
                else:
                    r_sum = float((be_sum/bt_sum)*100)
                write(5, bt_sum, be_sum)

            if (d_read == '번호'):
                ct_sum = total_sum(ct_sum)
                ce_sum = error_sum(ce_sum)
                if (ce_sum == 0):
                    r_sum = 0
                else:
                    r_sum = float((ce_sum/ct_sum)*100)
                write(8, ct_sum, ce_sum)

            if (d_read == '금액'):
                dt_sum = total_sum(dt_sum)
                de_sum = error_sum(de_sum)
                if (de_sum == 0):
                    r_sum = 0
                else:
                    r_sum = float((de_sum/dt_sum)*100)
                write(11, dt_sum, de_sum)

            if (d_read == '수량'):
                et_sum = total_sum(et_sum)
                ee_sum = error_sum(ee_sum)
                if (ee_sum == 0):
                    r_sum = 0
                else:
                    r_sum = float((ee_sum/et_sum)*100)
                write(14, et_sum, ee_sum)

            if (d_read == '율'):
                ft_sum = total_sum(ft_sum)
                fe_sum = error_sum(fe_sum)
                if (fe_sum == 0):
                    r_sum = 0
                else:
                    r_sum = float((fe_sum/ft_sum)*100)
                write(17, ft_sum, fe_sum)

            if (d_read == '코드'):
                gt_sum = total_sum(gt_sum)
                ge_sum = error_sum(ge_sum)
                if (ge_sum == 0):
                    r_sum = 0
                else:
                    r_sum = float((ge_sum/gt_sum)*100)
                write(20, gt_sum, ge_sum)

            i += 1

        output.cell(row=j, column=1, value='합계')

        for col in range (2, 23) :
            sum = 0.0
            for row in range (4, j) :
                if hasattr(output.cell(row=row, column=col), 'value') :
                    if output.cell(row=row, column=col).value is not None :
                        sum += output.cell(row=row, column=col).value
                    else :
                        output.cell(row=j, column=col, value = 0)
            output.cell(row=j, column=col, value=sum)

        f_name = cur_dir + '/test/output_' + f_name
        out.save(f_name)

        print('Done')
        check = input('continue? (y/n): ')
        if (check=='n') :
            break

